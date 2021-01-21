import pandas as pd
from tqdm import tqdm, trange
import openpyxl 

# 讀取stock_code內容
stock = openpyxl.load_workbook("./Result/2017_刪除無分析師預測後剩餘股吧名單.xlsx")
stock_sheet = stock['Sheet']

# 建立刪除股票名單
delete_stock = openpyxl.Workbook()
delete_stock_sheet = delete_stock['Sheet']

# 建立剩餘股票名單
new_stock = openpyxl.Workbook()
new_stock_sheet = new_stock['Sheet']

# 寫入第一列資訊
title = ["PostDate", "Stkcd", "TotalPosts", "AvgReadings", "AvgComments", "AvgNetComments", "AvgThumbUps", "AvgUserBarAge", "AvgUserInfluIndex", "AvgUserPosts", "AvgUserComments", "Indcd"]
for i in range(1, stock_sheet.max_column+1):
    new_stock_sheet.cell(row=1, column=i, value=title[i-1])
    delete_stock_sheet.cell(row=1, column=i, value=title[i-1])

# AvgUserInfluIndex column = 9
for i in trange(2, stock_sheet.max_row+1):

    # 確認列數
    delete_row = delete_stock_sheet.max_row+1
    new_row = new_stock_sheet.max_row+1

    if(stock_sheet.cell(row=i, column=9).value >= 2.5):
        # 不刪除的
        for j in range(1, new_stock_sheet.max_column+1):
            new_stock_sheet.cell(row=new_row, column=j, value=stock_sheet.cell(row=i, column=j).value)
    else:
        # 要刪除的
        for j in range(1, delete_stock_sheet.max_column+1):
            delete_stock_sheet.cell(row=delete_row, column=j, value=stock_sheet.cell(row=i, column=j).value)


# 計算刪除筆數
delete_stock_sheet.cell(row=1, column=delete_stock_sheet.max_column+2, value="刪除筆數")
delete_stock_sheet.cell(row=2, column=delete_stock_sheet.max_column, value=delete_stock_sheet.max_row-1)


# 存檔
delete_stock.save(r'./Result/2017_樣本小於2.5股吧名單.xlsx')
new_stock.save(r'./Result/2017_刪除樣本小於2.5後股吧名單.xlsx')

