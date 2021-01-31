import pandas as pd
from tqdm import tqdm, trange
import openpyxl 
import threading
import time

year = 2017
# for year in years:

# 讀取股吧內容
stock = openpyxl.load_workbook("./" + str(year) + "/" + str(year) + "股吧.xlsx")
stock_sheet = stock['工作表1']

# 讀取產業代碼
industrial_code = openpyxl.load_workbook("./" + str(year) + "/產業代碼" + str(year) + ".xlsx")
industrial_code_sheet = industrial_code['FI_T10']

# 欲刪除產業代碼
industrial_code_to_delete = ["J66", "J67", "J68", "J69"]

# 建立刪除股票名單
delete_stock = openpyxl.Workbook()
delete_stock_sheet = delete_stock['Sheet']

# 建立剩餘股票名單
new_stock = openpyxl.Workbook()
new_stock_sheet = new_stock['Sheet']

# 寫入第一列資訊
title = ["PostDate", "Stkcd", "TotalPosts", "AvgReadings", "AvgComments", "AvgNetComments", "AvgThumbUps", "AvgUserBarAge", "AvgUserInfluIndex", "AvgUserPosts", "AvgUserComments", "Indcd"]
new_stock_sheet.append(title)
delete_stock_sheet.append(title)

# 計算刪除股票號碼總數
delete_stock_stkcd = []

# 每1000筆
num = int((stock_sheet.max_row+1)/1000) 
start = [2,]
end = []

for i in range(num):
    end.append(start[i]+1000)
    start.append(end[i])

end.append(stock_sheet.max_row+1)
# print((industrial_code_sheet.cell(row=2, column=1).value))
# print("start = ", start )
# print("end = ", end)

# 計算刪除股票號碼總數
delete_stock_stkcd = []
print(num)
for nums in range(num+1):

    starting = int(start[nums])
    ending = int(end[nums])
    print(starting, ending)

    for i in trange(starting, ending):

        # 取得stkcd
        stock_stkcd = int(stock_sheet.cell(row=i, column=2).value)
        industrial_code_stkcd = int(industrial_code_sheet.cell(row=starting, column=1).value)

        # 使stkcd相等
        for k in range(2, industrial_code_sheet.max_row+1):
            industrial_code_stkcd = int(industrial_code_sheet.cell(row=starting, column=1).value)
            if (industrial_code_stkcd < stock_stkcd):  
                k = k+1
                starting = k
            else: break

        # 更新到該年最新的產業代碼
        for j in range(starting, industrial_code_sheet.max_row+1):
            if (industrial_code_sheet.cell(row=j, column=1).value != industrial_code_sheet.cell(row=j+1, column=1).value): 
                starting = j
                break


        # 取出產業代碼中的stkcd
        industrial_code_stkcd = int(industrial_code_sheet.cell(row=starting, column=1).value)

        # print(stock_stkcd, industrial_code_stkcd)
        # 寫入stkcd
        if(stock_stkcd==industrial_code_stkcd):
            # print("yes")
            # 取出indcd
            industrial_code_indcd = industrial_code_sheet.cell(row=starting, column=3).value
            
            # 複製原始檔案並寫入indcd
            stock_sheet.cell(row=i, column=12, value=industrial_code_indcd)
            
            # 確認列數
            delete_row = delete_stock_sheet.max_row+1
            new_row = new_stock_sheet.max_row+1       

            temp = int(stock_stkcd / (10**(len(str(stock_stkcd))-1)))
            
            # 寫入indcd
            if(industrial_code_indcd in industrial_code_to_delete):
                # 要刪除的
                for l in range(1, stock_sheet.max_column+1):
                    delete_stock_sheet.cell(row=delete_row, column=l, value=stock_sheet.cell(row=i, column=l).value)
                    if (stock_stkcd not in delete_stock_stkcd):
                        delete_stock_stkcd.append(stock_stkcd)
            elif( temp == 2 or temp == 9 ):
                # 要刪除的
                for l in range(1, stock_sheet.max_column+1):
                    delete_stock_sheet.cell(row=delete_row, column=l, value=stock_sheet.cell(row=i, column=l).value)
                    if (stock_stkcd not in delete_stock_stkcd):
                        delete_stock_stkcd.append(stock_stkcd)
            else:
                # 不刪除的
                for l in range(1, stock_sheet.max_column+1):
                    new_stock_sheet.cell(row=new_row, column=l, value=stock_sheet.cell(row=i, column=l).value)


# 最初 start = 2, end = stock_sheet.max_row+1
"""
def thread1(start, end, num):
    
    year = 2017
    # 讀取股吧內容
    stock = openpyxl.load_workbook("./" + str(year) + "/" + str(year) + "股吧.xlsx")
    stock_sheet = stock['工作表1']

    # 讀取產業代碼
    industrial_code = openpyxl.load_workbook("./" + str(year) + "/產業代碼" + str(year) + ".xlsx")
    industrial_code_sheet = industrial_code['FI_T10']
    
    # 欲刪除產業代碼
    industrial_code_to_delete = ["J66", "J67", "J68", "J69"]
    

    # 建立刪除股票名單
    delete_stock = openpyxl.Workbook()
    delete_stock_sheet = delete_stock['Sheet']

    # 建立剩餘股票名單
    new_stock = openpyxl.Workbook()
    new_stock_sheet = new_stock['Sheet']

    # 寫入第一列資訊
    title = ["PostDate", "Stkcd", "TotalPosts", "AvgReadings", "AvgComments", "AvgNetComments", "AvgThumbUps", "AvgUserBarAge", "AvgUserInfluIndex", "AvgUserPosts", "AvgUserComments", "Indcd"]
    new_stock_sheet.append(title)
    delete_stock_sheet.append(title)

    # 計算刪除股票號碼總數
    delete_stock_stkcd = []

    for i in range(start, end):

        # 取得stkcd
        stock_stkcd = int(stock_sheet.cell(row=i, column=2).value)
        industrial_code_stkcd = int(industrial_code_sheet.cell(row=start, column=1).value)

        # 使stkcd相等
        for k in range(1, industrial_code_sheet.max_row+1):
            industrial_code_stkcd = int(industrial_code_sheet.cell(row=start, column=1).value)
            if (industrial_code_stkcd < stock_stkcd):  
                k = k+1
                start = k
            else: break

        # 更新到該年最新的產業代碼
        for j in range(1, industrial_code_sheet.max_row+1):
            if (industrial_code_sheet.cell(row=j, column=1).value != industrial_code_sheet.cell(row=j+1, column=1).value): 
                start = j
                break

        # 取出產業代碼中的stkcd
        industrial_code_stkcd = int(industrial_code_sheet.cell(row=start, column=1).value)

        # 寫入stkcd
        if(stock_stkcd==industrial_code_stkcd):

            # 取出indcd
            industrial_code_indcd = industrial_code_sheet.cell(row=start, column=3).value
            
            # 複製原始檔案並寫入indcd
            stock_sheet.cell(row=i, column=12, value=industrial_code_indcd)
            
            # 確認列數
            delete_row = delete_stock_sheet.max_row+1
            new_row = new_stock_sheet.max_row+1       

            temp = int(stock_stkcd / (10**(len(str(stock_stkcd))-1)))

            # 寫入indcd
            if(industrial_code_indcd in industrial_code_to_delete):
                # 要刪除的
                for l in range(1, stock_sheet.max_column+1):
                    delete_stock_sheet.cell(row=delete_row, column=l, value=stock_sheet.cell(row=i, column=l).value)
                    if (stock_stkcd not in delete_stock_stkcd):
                        delete_stock_stkcd.append(stock_stkcd)
            elif( temp == 2 or temp == 9 ):
                # 要刪除的
                for l in range(1, stock_sheet.max_column+1):
                    delete_stock_sheet.cell(row=delete_row, column=l, value=stock_sheet.cell(row=i, column=l).value)
                    if (stock_stkcd not in delete_stock_stkcd):
                        delete_stock_stkcd.append(stock_stkcd)
            else:
                # 不刪除的
                for l in range(1, stock_sheet.max_column+1):
                    new_stock_sheet.cell(row=new_row, column=l, value=stock_sheet.cell(row=i, column=l).value)

    # 計算刪除筆數
    delete_stock_sheet.cell(row=1, column=delete_stock_sheet.max_column+2, value="刪除筆數總和")
    delete_stock_sheet.cell(row=2, column=delete_stock_sheet.max_column, value=delete_stock_sheet.max_row-1)

    # 計算刪除種類
    delete_stock_sheet.cell(row=1, column=delete_stock_sheet.max_column+2, value="刪除股票種類總和")
    delete_stock_sheet.cell(row=2, column=delete_stock_sheet.max_column, value=len(delete_stock_stkcd))

    # 存檔
    stock.save(r'./Result/' + str(year) + '/' + str(year) + '_加入產業代碼股吧' + str(num) + '.xlsx')
    delete_stock.save(r'./Result/' + str(year) + '/' + str(year) + '_金融產業股吧名單' + str(num) + '.xlsx')
    new_stock.save(r'./Result/' + str(year) + '/' + str(year) + '_刪除金融產業後剩餘股吧名單' + str(num) + '.xlsx')

num = int((stock_sheet.max_row+1)/10000) 

start = [2,]
end = []

for i in range(num):
    end.append(start[i]+999)
    start.append(end[i]+1)

end.append(stock_sheet.max_row+1)

threads = []
for i in trange(num):
    threads.append(threading.Thread(target = thread1, args = (start[i],end[i], i+1)))
    print(start[i], end[i])
    threads[i].start()


for i in trange(num):
    threads[i].join()
    print("Thread" , i, "finish")
"""


# 計算刪除筆數
delete_stock_sheet.cell(row=1, column=delete_stock_sheet.max_column+2, value="刪除筆數總和")
delete_stock_sheet.cell(row=2, column=delete_stock_sheet.max_column, value=delete_stock_sheet.max_row-1)

# 計算刪除種類
delete_stock_sheet.cell(row=1, column=delete_stock_sheet.max_column+2, value="刪除股票種類總和")
delete_stock_sheet.cell(row=2, column=delete_stock_sheet.max_column, value=len(delete_stock_stkcd))

# 存檔
stock.save(r'./Result/' + str(year) + '/' + str(year) + '_加入產業代碼股吧.xlsx')
delete_stock.save(r'./Result/' + str(year) + '/' + str(year) + '_金融產業股吧名單.xlsx')
new_stock.save(r'./Result/' + str(year) + '/' + str(year) + '_刪除金融產業後剩餘股吧名單.xlsx')

print("finish")
