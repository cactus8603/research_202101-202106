import pandas as pd
import openpyxl 

years = [2017]
"""
for year in years:
    # print(year)
    # stock = openpyxl.load_workbook("./" + str(year) + "/" + str(year) + "股吧.xlsx")
    stock = openpyxl.load_workbook("./" + str(year) + "/產業代碼" + str(year) + ".xlsx")
    stock_sheet = stock['FI_T10']
    temp = str(stock_sheet.cell(row=2 , column=3).value)
    print(temp.isalnum())
    temp1 = int(temp)
    print(temp1%len(temp))
    # print(year)

stock_stkcd = 2999
print(10**len(str(stock_stkcd)))
print( int(stock_stkcd / (10**(len(str(stock_stkcd))-1))) )

stock = openpyxl.load_workbook("./" + str(years[0]) + "/" + str(years[0]) + "股吧.xlsx")
stock_sheet = stock['工作表1']

print(int(stock_sheet.cell(row=2 , column=2).value))
"""

import threading
import time

# 子執行緒的工作函數
def job(num):
  print("Thread", num)
  time.sleep(1)

# 建立 5 個子執行緒
threads = []
for i in range(5):
  threads.append(threading.Thread(target = job, args = (i,)))
  threads[i].start()

# 主執行緒繼續執行自己的工作
# ...

# 等待所有子執行緒結束
for i in range(5):
  threads[i].join()
  print(i)

print("Done.")
"""

c = 9943

num = int((c+1)/1000) 

start = [2,]
end = []
print("num =", num )

for i in range(num):
    end.append(start[i]+999)
    start.append(end[i]+1)

end.append(c+1)

print("start = ", start )
print("end = ", end)
"""
        str = (industrial_code_sheet.cell(row=starting, column=1).value)
        if (str.isalpha()):
            industrial_code_stkcd = int(industrial_code_sheet.cell(row=starting, column=1).value)
        else : industrial_code_stkcd = (industrial_code_sheet.cell(row=starting, column=1).value)
