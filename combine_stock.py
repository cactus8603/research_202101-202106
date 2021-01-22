import pandas as pd
from tqdm import tqdm, trange
import openpyxl 

# 建立新股票名單
new_stock = openpyxl.Workbook()
new_stock_sheet = new_stock['Sheet']

# 寫入第一列資訊
title = ["PostDate", "Stkcd", "TotalPosts", "AvgReadings", "AvgComments", "AvgNetComments", "AvgThumbUps", "AvgUserBarAge", "AvgUserInfluIndex", "AvgUserPosts", "AvgUserComments", "Indcd"]
new_stock_sheet.append(title)

index = [1, 2]
year = 2018
"""
# 讀取股吧內容
stock = openpyxl.load_workbook("./" + str(year) + "/" + str(year) + str(index[0]) + "股吧.xlsx")
stock_sheet = stock['工作表1']

# print(tuple(stock_sheet[1][1]).value)
print(stock_sheet[2][0].value)
"""
for ind in index:
    # 讀取股吧內容
    stock = openpyxl.load_workbook("./" + str(year) + "/" + str(year) + str(ind) + "股吧.xlsx")
    stock_sheet = stock['工作表1']

    row_temp = 2
    
    # stock_sheet.max_row+1
    for i in trange(2, 30):
        new_row = new_stock_sheet.max_row+1
        for j in range(1, stock_sheet.max_column+1):
            new_stock_sheet.cell(row=new_row, column=j, value=stock_sheet.cell(row=i, column=j).value)

new_stock.save(r'./' + str(year) + '/' + str(year) + '股吧.xlsx')
