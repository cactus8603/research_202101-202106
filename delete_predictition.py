import pandas as pd
from tqdm import tqdm, trange
import openpyxl 

# years = ["2017", "2018", "2019"]
years = ["2017"]
for year in years:
    """
    # 讀取stock_code內容
    stock = openpyxl.load_workbook("./Result/" + str(year) + "/" + str(year) + "_加入產業代碼股吧.xlsx")
    stock_sheet = stock['工作表1']
    """

    # 讀取stock_code內容
    stock = openpyxl.load_workbook("./Result/" + str(year) + "/" + str(year) + "_刪除金融產業後剩餘股吧名單.xlsx")
    stock_sheet = stock['Sheet']


    # 讀取分析師預測內容
    predict_stock = openpyxl.load_workbook("./" + str(year) + "/分析師預測" + str(year) + ".xlsx")
    predict_stock_sheet = predict_stock['工作表1']

    # 建立刪除股票名單
    delete_stock = openpyxl.Workbook()
    delete_stock_sheet = delete_stock['Sheet']

    # 建立剩餘股票名單
    new_stock = openpyxl.Workbook()
    new_stock_sheet = new_stock['Sheet']

    # 寫入第一列資訊
    title = ["PostDate", "Stkcd", "TotalPosts", "AvgReadings", "AvgComments", "AvgNetComments", "AvgThumbUps", "AvgUserBarAge", "AvgUserInfluIndex", "AvgUserPosts", "AvgUserComments", "Indcd"]
    new_stock_sheet.append(title)
    delete_stock_sheet.append(title)

    # 第二列開始讀入資料
    predict_stkcd = []
    # predict_stock_sheet.max_row+1
    for i in trange(2, predict_stock_sheet.max_row+1):
        if (predict_stock_sheet.cell(row=i, column=1).value != predict_stock_sheet.cell(row=i+1, column=1).value): 
            predict_stkcd.append(int(predict_stock_sheet.cell(row=i, column=1).value))
            # print((predict_stock_sheet.cell(row=i, column=1).value))

    # print(len(predict_stkcd))
    # stock_sheet.max_row+1
    for i in trange(2, stock_sheet.max_row+1):

        # 確認列數
        delete_row = delete_stock_sheet.max_row+1
        new_row = new_stock_sheet.max_row+1

        # 寫入
        if (int(stock_sheet.cell(row=i, column=2).value) in predict_stkcd):
            # 不刪除的
            for j in range(1, new_stock_sheet.max_column+1):
                new_stock_sheet.cell(row=new_row, column=j, value=stock_sheet.cell(row=i, column=j).value)
            
        else:
            # 要刪除的
            for j in range(1, delete_stock_sheet.max_column+1):
                delete_stock_sheet.cell(row=delete_row, column=j, value=stock_sheet.cell(row=i, column=j).value)
            
            # print(stock_sheet.cell(row=i, column=2).value)

    # 計算刪除筆數
    delete_stock_sheet.cell(row=1, column=delete_stock_sheet.max_column+2, value="刪除筆數")
    delete_stock_sheet.cell(row=2, column=delete_stock_sheet.max_column, value=delete_stock_sheet.max_row-1)

    # 存檔
    delete_stock.save(r'./Result/' + str(year) + '/' + str(year) + '_無分析師預測股吧名單.xlsx')
    new_stock.save(r'./Result/' + str(year) + '/' + str(year) + '_刪除無分析師預測後剩餘股吧名單.xlsx')

